import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
from reportlab.pdfgen import canvas

def convert_excel_to_pdf(excel_file, pdf_file):
    # Load the Excel workbook
    wb = load_workbook(excel_file)

    # Create a PDF canvas
    pdf = canvas.Canvas(pdf_file)

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        # Select the current sheet
        sheet = wb[sheet_name]

        # Iterate through all rows and columns in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                # Get the value of the cell
                cell_value = cell.value

                # Draw the cell value on the PDF canvas
                pdf.drawString(cell.column * 50, (sheet.max_row - cell.row + 1) * 20, str(cell_value))

    # Save the PDF file
    pdf.save()

# Example usage
excel_file_path = 'example.xlsx'
pdf_file_path = 'output.pdf'
convert_excel_to_pdf(excel_file_path, pdf_file_path)
